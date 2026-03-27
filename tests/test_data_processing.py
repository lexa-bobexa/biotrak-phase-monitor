from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from logic.data_processing import (
    _format_ctgov_phase,
    create_results_workbook_bytes,
    get_trials,
)


def test_create_results_workbook_bytes_deduplicates_nct_numbers():
    dataframe = pd.DataFrame(
        [
            {"NCT Number": "NCT001", "Product Name": "Drug A"},
            {"NCT Number": "NCT001", "Product Name": "Drug A"},
        ]
    )
    workbook_bytes = create_results_workbook_bytes({"SheetA": dataframe}, ["SheetA"])
    loaded = pd.read_excel(BytesIO(workbook_bytes), sheet_name=None)

    assert "SheetA" in loaded
    assert len(loaded["SheetA"]) == 1
    assert loaded["SheetA"].iloc[0]["NCT Number"] == "NCT001"


def test_get_trials_returns_filtered_trial_and_reports_progress(monkeypatch):
    input_frame = pd.DataFrame(
        [
            {
                "bioTRAK Product ID": "A-1",
                "Product Name": "Drug A",
                "Original Phase": "Phase 1",
            }
        ]
    )
    progress_updates = []

    class MockResponse:
        status_code = 200

        @staticmethod
        def json():
            return {
                "nextPageToken": None,
                "studies": [
                    {
                        "protocolSection": {
                            "identificationModule": {"nctId": "NCT001"},
                            "statusModule": {
                                "overallStatus": "RECRUITING",
                                "whyStopped": "Funding withdrawn",
                                "startDateStruct": {"date": "2025-01-01"},
                                "completionDateStruct": {"date": "2026-01-01"},
                            },
                            "sponsorCollaboratorsModule": {"leadSponsor": {"name": "Sponsor A"}},
                            "designModule": {"phases": ["PHASE2"]},
                            "oversightModule": {"isFdaRegulatedDrug": True},
                            "conditionsModule": {"conditions": ["Condition A"]},
                            "armsInterventionsModule": {"interventions": [{"name": "Drug A"}]},
                            "contactsLocationsModule": {
                                "locations": [{"country": "United States"}]
                            },
                        }
                    }
                ],
            }

    def mock_get(_url, timeout):
        assert timeout == 30
        return MockResponse()

    monkeypatch.setattr("logic.data_processing.requests.get", mock_get)

    results = get_trials(
        input_frame,
        "bioTRAK Product ID",
        progress_callback=lambda pct, msg: progress_updates.append((pct, msg)),
    )

    assert len(results) == 1
    assert results[0]["NCT Number"] == "NCT001"
    assert results[0]["Phase on CT.gov"] == "Phase 2"
    assert results[0]["Why Stopped"] == "Funding withdrawn"
    assert progress_updates[-1][0] == 100


def test_format_ctgov_phase_includes_multiple_phases():
    phase_value = _format_ctgov_phase(["PHASE1", "PHASE2"])
    assert phase_value == "Phase 1/Phase 2"


def test_get_trials_defaults_why_stopped_when_missing(monkeypatch):
    input_frame = pd.DataFrame(
        [
            {
                "bioTRAK Product ID": "A-1",
                "Product Name": "Drug A",
                "Original Phase": "Phase 1",
            }
        ]
    )

    class MockResponse:
        status_code = 200

        @staticmethod
        def json():
            return {
                "nextPageToken": None,
                "studies": [
                    {
                        "protocolSection": {
                            "identificationModule": {"nctId": "NCT002"},
                            "statusModule": {"overallStatus": "COMPLETED"},
                            "sponsorCollaboratorsModule": {"leadSponsor": {"name": "Sponsor B"}},
                            "designModule": {"phases": ["PHASE1"]},
                            "oversightModule": {"isFdaRegulatedDrug": False},
                            "conditionsModule": {"conditions": []},
                            "armsInterventionsModule": {"interventions": [{"name": "Drug A"}]},
                            "contactsLocationsModule": {
                                "locations": [{"country": "United States"}]
                            },
                        }
                    }
                ],
            }

    monkeypatch.setattr("logic.data_processing.requests.get", lambda _url, timeout: MockResponse())

    results = get_trials(input_frame, "bioTRAK Product ID")

    assert len(results) == 1
    assert results[0]["NCT Number"] == "NCT002"
    assert results[0]["Why Stopped"] == "Not Available"


def test_create_results_workbook_bytes_sets_nct_number_hyperlink():
    dataframe = pd.DataFrame(
        [
            {
                "NCT Number": "NCT777",
            }
        ]
    )
    workbook_bytes = create_results_workbook_bytes({"SheetA": dataframe}, ["SheetA"])

    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook["SheetA"]
    headers = [cell.value for cell in worksheet[1]]
    nct_column = headers.index("NCT Number") + 1
    nct_cell = worksheet.cell(row=2, column=nct_column)

    assert nct_cell.value == "NCT777"
    assert nct_cell.hyperlink is not None
    assert nct_cell.hyperlink.target == "https://clinicaltrials.gov/study/NCT777"


def test_get_trials_excludes_studies_older_than_eight_year_cutoff(monkeypatch):
    input_frame = pd.DataFrame(
        [
            {
                "bioTRAK Product ID": "A-1",
                "Product Name": "Drug A",
                "Original Phase": "Phase 1",
            }
        ]
    )

    class MockResponse:
        status_code = 200

        @staticmethod
        def json():
            return {
                "nextPageToken": None,
                "studies": [
                    {
                        "protocolSection": {
                            "identificationModule": {"nctId": "NCT_OLD"},
                            "statusModule": {
                                "overallStatus": "COMPLETED",
                                "completionDateStruct": {"date": "2010-01-01"},
                            },
                            "sponsorCollaboratorsModule": {"leadSponsor": {"name": "Sponsor Old"}},
                            "designModule": {"phases": ["PHASE1"]},
                            "oversightModule": {"isFdaRegulatedDrug": False},
                            "conditionsModule": {"conditions": []},
                            "armsInterventionsModule": {"interventions": [{"name": "Drug A"}]},
                            "contactsLocationsModule": {
                                "locations": [{"country": "United States"}]
                            },
                        }
                    },
                    {
                        "protocolSection": {
                            "identificationModule": {"nctId": "NCT_RECENT"},
                            "statusModule": {
                                "overallStatus": "COMPLETED",
                                "completionDateStruct": {"date": "2024-01-01"},
                            },
                            "sponsorCollaboratorsModule": {"leadSponsor": {"name": "Sponsor New"}},
                            "designModule": {"phases": ["PHASE2"]},
                            "oversightModule": {"isFdaRegulatedDrug": True},
                            "conditionsModule": {"conditions": ["Condition A"]},
                            "armsInterventionsModule": {"interventions": [{"name": "Drug A"}]},
                            "contactsLocationsModule": {
                                "locations": [{"country": "United States"}]
                            },
                        }
                    },
                ],
            }

    monkeypatch.setattr("logic.data_processing.requests.get", lambda _url, timeout: MockResponse())

    results = get_trials(
        input_frame,
        "bioTRAK Product ID",
        trial_end_cutoff_years=8,
        reference_timestamp=pd.Timestamp("2026-03-27"),
    )

    assert len(results) == 1
    assert results[0]["NCT Number"] == "NCT_RECENT"
